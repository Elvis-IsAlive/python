import openpyxl

wb = openpyxl.load_workbook("einwohner_Bayern.xlsx")
ws = wb.active

data = {}

for i in range(2, ws.max_row):
    stadt = ws.cell(i, 1).value
    bezirk = ws.cell(i, 3).value
    einwohner = ws.cell(i,4).value

    data.setdefault(bezirk, {})
    data[bezirk].setdefault(stadt, {})

    data[bezirk][stadt]["Einwhoner"] = einwohner
    
    
