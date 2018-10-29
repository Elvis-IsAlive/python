#! /usr/bin/python3.6

import openpyxl, sys
from openpyxl.styles import Font

filename = "multiplicationTable.xlsx"

num = int(sys.argv[1])


print("number : %i" % num)

wb = openpyxl.Workbook()
s = wb.active


for i in range(1, num + 2):
    for j in range(1, num + 2):
        if i == 1 and j == 1:
            continue
        elif i == 1:
            s.cell(i, j).value = j - 1
        elif j == 1:
            s.cell(i, j).value = i - 1
        else:
            s.cell(i, j).value = (i - 1) * (j - 1)  

# Formatierung
for c in s["1:1"]:
    c.font = Font(bold = True)
    
for c in s["A:A"]:
    c.font = Font(bold = True)
    


wb.save(filename)
